import os
import sys
import openpyxl
import requests

DATA_DIR = '/Users/joseph-macbookpro-m3/Documents/Apps/practice/data'
ENV_FILE = os.path.join(os.path.dirname(__file__), '..', '.env.local')
ORG_ID = '00000000-0000-0000-0000-000000000001'

# Load .env.local manually (avoid python-dotenv dependency)
env = {}
with open(ENV_FILE) as f:
    for line in f:
        line = line.strip()
        if line and not line.startswith('#') and '=' in line:
            k, v = line.split('=', 1)
            env[k.strip()] = v.strip()

SUPABASE_URL = env['NEXT_PUBLIC_SUPABASE_URL']
SERVICE_KEY = env['SUPABASE_SERVICE_ROLE_KEY']

BASE_HEADERS = {
    'apikey': SERVICE_KEY,
    'Authorization': f'Bearer {SERVICE_KEY}',
    'Content-Type': 'application/json',
}


def supabase_get(table, params=None):
    url = f'{SUPABASE_URL}/rest/v1/{table}'
    r = requests.get(url, headers=BASE_HEADERS, params=params)
    if not r.ok:
        print(f'  ERROR GET {table}: {r.status_code} {r.text[:300]}')
        sys.exit(1)
    return r.json()


def supabase_delete(table):
    url = f'{SUPABASE_URL}/rest/v1/{table}?id=not.is.null'
    r = requests.delete(url, headers={**BASE_HEADERS, 'Prefer': 'return=minimal'})
    if r.ok:
        print(f'  Cleared {table}')
    elif r.status_code in (404, 400) and ('not found' in r.text.lower() or 'does not exist' in r.text.lower()):
        print(f'  Skipped {table} (table not found)')
    else:
        print(f'  ERROR deleting {table}: {r.status_code} {r.text[:200]}')


def supabase_insert(table, rows, return_rep=True):
    url = f'{SUPABASE_URL}/rest/v1/{table}'
    prefer = 'return=representation' if return_rep else 'return=minimal'
    r = requests.post(url, headers={**BASE_HEADERS, 'Prefer': prefer}, json=rows)
    if not r.ok:
        print(f'  ERROR POST {table}: {r.status_code} {r.text[:400]}')
        sys.exit(1)
    return r.json() if return_rep else []


def batch_insert(table, rows, batch_size=200, return_rep=True):
    results = []
    total = len(rows)
    for i in range(0, total, batch_size):
        batch = rows[i:i + batch_size]
        batch_results = supabase_insert(table, batch, return_rep=return_rep)
        if return_rep:
            results.extend(batch_results)
        done = min(i + batch_size, total)
        print(f'  {table}: {done}/{total}')
    return results


# ─── STEP 1: Clear data ───────────────────────────────────────────────────────
print('\n=== STEP 1: Clearing data (preserving orgs/branches/profiles) ===')

tables_to_clear = [
    'transaction_item_supplier_costs',
    'transaction_items',
    'installment_plans',
    'accounts_receivable',
    'transactions',
    'quotation_items',
    'quotations',
    'stock_transfer_items',
    'branch_stock_request_items',
    'branch_stock_requests',
    'stock_transfers',
    'purchase_order_items',
    'purchase_order_cheques',
    'purchase_orders',
    'inventory_movements',
    'inventory',
    'bundle_items',
    'bundles',
    'product_supplier_stock',
    'product_suppliers',
    'products',
    'categories',
    'suppliers',
    'customers',
    'expenses',
    'error_logs',
]
for t in tables_to_clear:
    supabase_delete(t)


# ─── STEP 2: Categories ───────────────────────────────────────────────────────
print('\n=== STEP 2: Inserting categories ===')
wb = openpyxl.load_workbook(f'{DATA_DIR}/CATEGORY.xlsx')
ws = wb.active
col_headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
code_col = col_headers.index('Code')
name_col = col_headers.index('Name')

cat_rows = []
code_to_name = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    code = row[code_col]
    name = row[name_col]
    if name:
        name_str = str(name).strip()
        cat_rows.append({'org_id': ORG_ID, 'name': name_str})
        code_to_name[code] = name_str

inserted_cats = supabase_insert('categories', cat_rows)
print(f'  Inserted {len(inserted_cats)} categories')
name_to_uuid = {r['name']: r['id'] for r in inserted_cats}
code_to_uuid = {code: name_to_uuid.get(name) for code, name in code_to_name.items()}


# ─── STEP 3: Suppliers ────────────────────────────────────────────────────────
print('\n=== STEP 3: Inserting suppliers ===')
wb2 = openpyxl.load_workbook(f'{DATA_DIR}/supplier.xlsx')
ws2 = wb2.active
h2 = [c.value for c in next(ws2.iter_rows(min_row=1, max_row=1))]
sup_name_col = h2.index('Name')
sup_phone_col = h2.index('ContactNumber')
sup_contact_col = h2.index('Contact')

sup_rows = []
for row in ws2.iter_rows(min_row=2, values_only=True):
    name = row[sup_name_col]
    if not name:
        continue
    phone = row[sup_phone_col] if sup_phone_col < len(row) else None
    contact = row[sup_contact_col] if sup_contact_col < len(row) else None
    sup_rows.append({
        'org_id': ORG_ID,
        'name': str(name).strip(),
        'phone': str(phone).strip() if phone else None,
        'contact_name': str(contact).strip() if contact else None,
    })

inserted_sups = supabase_insert('suppliers', sup_rows)
print(f'  Inserted {len(inserted_sups)} suppliers')


# ─── STEP 4: Products ─────────────────────────────────────────────────────────
print('\n=== STEP 4: Inserting products ===')
wb3 = openpyxl.load_workbook(f'{DATA_DIR}/ITEM.xlsx')
ws3 = wb3.active
h3 = [c.value for c in next(ws3.iter_rows(min_row=1, max_row=1))]
bc_col = h3.index('Barcode')
pname_col = h3.index('name')
cost_col = h3.index('UnitCost')
price_col = h3.index('UnitPrice')
catcode_col = h3.index('CategoryCode')

prod_rows = []
seen_skus: dict[str, int] = {}
for row in ws3.iter_rows(min_row=2, values_only=True):
    name = row[pname_col]
    if not name:
        continue
    barcode = row[bc_col]
    barcode_str = str(barcode).strip() if barcode is not None else None
    # strip trailing .0 from numeric barcodes read as float
    if barcode_str and barcode_str.endswith('.0') and barcode_str[:-2].isdigit():
        barcode_str = barcode_str[:-2]
    sku_base = barcode_str if barcode_str else f'ITEM-{len(prod_rows)+1}'
    # deduplicate SKUs
    if sku_base in seen_skus:
        seen_skus[sku_base] += 1
        sku = f'{sku_base}-{seen_skus[sku_base]}'
    else:
        seen_skus[sku_base] = 1
        sku = sku_base

    cat_code = row[catcode_col]
    cat_uuid = code_to_uuid.get(cat_code) if cat_code is not None else None
    cost = row[cost_col]
    price = row[price_col]

    prod_rows.append({
        'org_id': ORG_ID,
        'sku': sku,
        'barcode': barcode_str,
        'name': str(name).strip(),
        'cost_price': float(cost) if cost is not None else 0.0,
        'selling_price': float(price) if price is not None else 0.0,
        'category_id': cat_uuid,
        'unit': 'pcs',
        'is_active': True,
    })

print(f'  {len(prod_rows)} products to insert')
all_products = batch_insert('products', prod_rows, batch_size=200)
print(f'  Total products inserted: {len(all_products)}')


# ─── STEP 5: Fetch branches ───────────────────────────────────────────────────
print('\n=== STEP 5: Fetching branches ===')
branches = supabase_get('branches', params={'org_id': f'eq.{ORG_ID}', 'select': 'id,name'})
print(f'  Found {len(branches)} branch(es): {[b["name"] for b in branches]}')


# ─── STEP 6: Inventory rows ───────────────────────────────────────────────────
print('\n=== STEP 6: Inserting inventory ===')
inv_rows = [
    {'product_id': p['id'], 'branch_id': b['id'], 'quantity': 0, 'low_stock_threshold': 10}
    for p in all_products
    for b in branches
]
print(f'  {len(inv_rows)} inventory rows to insert')
batch_insert('inventory', inv_rows, batch_size=500, return_rep=False)


# ─── Summary ──────────────────────────────────────────────────────────────────
print('\n=== DONE ===')
print(f'  Categories : {len(cat_rows)}')
print(f'  Suppliers  : {len(sup_rows)}')
print(f'  Products   : {len(all_products)}')
print(f'  Inventory  : {len(inv_rows)} rows ({len(branches)} branch(es) × {len(all_products)} products)')
